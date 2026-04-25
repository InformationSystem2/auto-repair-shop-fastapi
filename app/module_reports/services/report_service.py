import csv
import io
from typing import Any


def generate_csv(
    columns: list[str], column_labels: dict[str, str], rows: list[dict[str, Any]]
) -> bytes:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=columns)
    writer.writerow({c: column_labels.get(c, c) for c in columns})
    for row in rows:
        writer.writerow({c: row.get(c, "") for c in columns})
    return output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


def generate_excel(
    columns: list[str], column_labels: dict[str, str], rows: list[dict[str, Any]],
    sheet_name: str = "Reporte",
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name  # type: ignore[union-attr]

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_key in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column_labels.get(col_key, col_key))  # type: ignore[union-attr]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18  # type: ignore[union-attr]

    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_key in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(row.get(col_key, "") or ""))  # type: ignore[union-attr]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_html(
    columns: list[str],
    column_labels: dict[str, str],
    rows: list[dict[str, Any]],
    title: str = "Reporte",
    total_label: str = "Total de registros",
) -> bytes:
    headers = "".join(f"<th>{column_labels.get(c, c)}</th>" for c in columns)
    body_rows = ""
    for row in rows:
        cells = "".join(f"<td>{row.get(c, '') or ''}</td>" for c in columns)
        body_rows += f"<tr>{cells}</tr>"

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>{title}</title>
<style>
  body {{ font-family: Arial, sans-serif; padding: 20px; }}
  h1 {{ color: #4F46E5; }}
  table {{ border-collapse: collapse; width: 100%; margin-top: 16px; }}
  th {{ background: #4F46E5; color: #fff; padding: 10px 12px; text-align: left; font-size: 13px; }}
  td {{ border-bottom: 1px solid #e5e7eb; padding: 8px 12px; font-size: 13px; }}
  tr:hover td {{ background: #f9fafb; }}
</style>
</head>
<body>
<h1>{title}</h1>
<p>{total_label}: {len(rows)}</p>
<table>
  <thead><tr>{headers}</tr></thead>
  <tbody>{body_rows}</tbody>
</table>
</body>
</html>"""
    return html.encode("utf-8")


def generate_pdf(
    columns: list[str],
    column_labels: dict[str, str],
    rows: list[dict[str, Any]],
    title: str = "Reporte",
    total_label: str = "Total de registros",
) -> bytes:
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=30,
        bottomMargin=20,
    )
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(title, styles["Title"]),
        Spacer(1, 12),
        Paragraph(f"{total_label}: {len(rows)}", styles["Normal"]),
        Spacer(1, 12),
    ]

    header_row = [column_labels.get(c, c) for c in columns]
    data = [header_row]
    for row in rows:
        data.append([str(row.get(c, "") or "") for c in columns])

    col_count = len(columns)
    page_width = landscape(A4)[0] - 40
    col_width = page_width / col_count if col_count else page_width

    table = Table(data, colWidths=[col_width] * col_count, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E7EB")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()
